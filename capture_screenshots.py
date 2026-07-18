"""
Render cropped screenshots of key workbook sections for the README.
Uses Pillow to draw clean table images — no external renderer needed.
"""
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import textwrap, os, math

WB_PATH = r"D:\ISKORE\iskolar-tracker.xlsx"
OUT_DIR = r"D:\ISKORE\screenshots"
os.makedirs(OUT_DIR, exist_ok=True)

# ── Colours ─────────────────────────────────────────────────────────────
NAVY       = (31, 78, 121)
WHITE      = (255, 255, 255)
LIGHT_BG   = (242, 247, 251)
DARK_TEXT   = (30, 30, 30)
GRAY_TEXT   = (120, 120, 120)
GREEN_BG   = (198, 239, 206)
GREEN_FG   = (0, 97, 0)
RED_BG     = (255, 199, 206)
RED_FG     = (156, 0, 6)
YELLOW_BG  = (255, 235, 156)
YELLOW_FG  = (156, 101, 0)
BORDER_CLR = (176, 176, 176)

# ── Fonts ───────────────────────────────────────────────────────────────
try:
    FONT_BOLD = ImageFont.truetype("arialbd.ttf", 16)
    FONT_BOLD_SM = ImageFont.truetype("arialbd.ttf", 13)
    FONT = ImageFont.truetype("arial.ttf", 14)
    FONT_SM = ImageFont.truetype("arial.ttf", 12)
    FONT_HDR = ImageFont.truetype("arialbd.ttf", 18)
except:
    FONT_BOLD = ImageFont.load_default()
    FONT_BOLD_SM = ImageFont.load_default()
    FONT = ImageFont.load_default()
    FONT_SM = ImageFont.load_default()
    FONT_HDR = ImageFont.load_default()

def text_size(draw, text, font):
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]

def wrap_text(text, max_w, draw, font):
    """Word-wrap text to fit within max_w pixels."""
    if not text:
        return [""]
    w, _ = text_size(draw, text, font)
    if w <= max_w:
        return [text]
    words = text.split(" ")
    lines = []
    cur = ""
    for word in words:
        test = cur + " " + word if cur else word
        tw, _ = text_size(draw, test, font)
        if tw <= max_w:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = word
    if cur:
        lines.append(cur)
    return lines if lines else [text[:int(len(text)*max_w/w)]]

def truncate(text, max_w, draw, font):
    w, _ = text_size(draw, text, font)
    if w <= max_w:
        return text
    while text:
        text = text[:-1]
        w, _ = text_size(draw, text + "...", font)
        if w <= max_w:
            return text + "..."
    return ""

def render_table(title, headers, rows, col_widths, row_height=42,
                 header_h=44, title_h=52, pad_l=12, max_rows=25):
    """
    Draw a table as a PIL Image.
    headers: list of strings
    rows: list of lists of (value, bg_color, fg_color, font_override)
    col_widths: list of pixel widths
    """
    ncols = len(headers)
    total_w = sum(col_widths) + (ncols + 1) * 1 + pad_l * 2
    nrows = min(len(rows), max_rows)
    total_h = title_h + header_h + nrows * row_height + 4

    img = Image.new("RGB", (total_w, total_h), WHITE)
    draw = ImageDraw.Draw(img)

    y = 0
    # ── Title ──
    draw.text((pad_l, 6), title, fill=NAVY, font=FONT_HDR)
    y += title_h

    # ── Header row ──
    x = pad_l
    for ci, hdr in enumerate(headers):
        cw = col_widths[ci]
        draw.rectangle([x, y, x + cw, y + header_h], fill=NAVY)
        draw.text((x + 4, y + (header_h - text_size(draw, hdr, FONT_BOLD_SM)[1]) // 2),
                  hdr, fill=WHITE, font=FONT_BOLD_SM)
        x += cw + 1
    y += header_h

    # ── Data rows ──
    for ri in range(nrows):
        x = pad_l
        bg = LIGHT_BG if ri % 2 == 0 else WHITE
        row_data = rows[ri]
        for ci in range(ncols):
            cw = col_widths[ci]
            val, cell_bg, cell_fg, font_ovr = row_data[ci] if ci < len(row_data) else ("", None, None, None)
            use_bg = cell_bg or bg
            use_fg = cell_fg or DARK_TEXT
            use_font = font_ovr or FONT_SM
            # Draw cell background
            draw.rectangle([x, y, x + cw, y + row_height], fill=use_bg)
            # Draw text
            if val:
                display = str(val)
                tw, _ = text_size(draw, display, use_font)
                if tw > cw - 8:
                    display = truncate(display, cw - 8, draw, use_font)
                draw.text((x + 4, y + (row_height - text_size(draw, display, use_font)[1]) // 2),
                          display, fill=use_fg, font=use_font)
            x += cw + 1
        y += row_height

    return img


# ══════════════════════════════════════════════════════════════════════
# Screenshot 1 — Dashboard (Held Awards + Portfolio Summary)
# ══════════════════════════════════════════════════════════════════════
def capture_dashboard():
    wb = load_workbook(WB_PATH)
    ws = wb["Dashboard"]

    # Held awards section (rows 6-7 in sheet, 0-indexed 5-6)
    held_rows = []
    for r in range(6, 8):
        a = ws.cell(r, 1).value or ""
        b = ws.cell(r, 2).value or ""
        c = ws.cell(r, 3).value or ""
        held_rows.append([
            (a, None, DARK_TEXT, FONT_SM),
            (b, None, DARK_TEXT, FONT_SM),
            (c, GREEN_BG, GREEN_FG, FONT_BOLD_SM),
        ])

    # Portfolio summary (rows 10-16)
    stats_rows = []
    for r in range(10, 17):
        a = ws.cell(r, 1).value or ""
        b = ws.cell(r, 2).value or ""
        # Determine color
        bg = None
        if "Eligible" in str(a):
            bg = GREEN_BG
        elif "Ineligible" in str(a):
            bg = RED_BG
        elif "Conditional" in str(a):
            bg = YELLOW_BG
        elif "Open" in str(a):
            bg = (226, 239, 218)

        stats_rows.append([
            (str(a), None, DARK_TEXT, FONT_SM),
            (str(b), bg, NAVY, FONT_BOLD),
        ])

    # Combine into a single image with two sections
    left_w = 520
    right_w = 140
    total_w = left_w + right_w + 40
    row_h = 38
    total_h = 60 + (len(held_rows) + len(stats_rows) + 2) * row_h + 30

    img = Image.new("RGB", (total_w, total_h), WHITE)
    draw = ImageDraw.Draw(img)

    y = 8
    # Section: Currently Held
    draw.text((12, y), "CURRENTLY HELD AWARDS", fill=NAVY, font=FONT_BOLD)
    y += 38
    # mini header
    for ci, (h, cw) in enumerate([
        ("Award", 280), ("Organization", 180), ("Status", 100)
    ]):
        draw.rectangle([12 + sum([0, 280, 180][:ci]), y,
                        12 + sum([0, 280, 180][:ci]) + cw, y + 30], fill=NAVY)
        draw.text((16 + sum([0, 280, 180][:ci]), y + 4), h, fill=WHITE, font=FONT_BOLD_SM)
    y += 32

    for row in held_rows:
        for ci, (val, bg_clr, fg, fnt) in enumerate(row):
            offsets = [0, 280, 180]
            cw = [280, 180, 100][ci]
            bx = 12 + offsets[ci]
            draw.rectangle([bx, y, bx + cw, y + row_h], fill=bg_clr or WHITE)
            draw.text((bx + 4, y + (row_h - text_size(draw, val, fnt)[1]) // 2),
                      val, fill=fg, font=fnt)
        y += row_h

    y += 12
    # Section: Portfolio Summary
    draw.text((12, y), "PORTFOLIO SUMMARY", fill=NAVY, font=FONT_BOLD)
    y += 38

    for row in stats_rows:
        for ci, (val, bg_clr, fg, fnt) in enumerate(row):
            cw = [left_w, right_w][ci]
            bx = 12 + [0, left_w][ci]
            draw.rectangle([bx, y, bx + cw, y + row_h], fill=bg_clr or WHITE)
            tw, _ = text_size(draw, val, fnt)
            if ci == 1:
                tx = bx + (cw - tw) // 2
            else:
                tx = bx + 6
            draw.text((tx, y + (row_h - text_size(draw, val, fnt)[1]) // 2),
                      val, fill=fg, font=fnt)
        y += row_h

    img.save(os.path.join(OUT_DIR, "01-dashboard.png"))
    wb.close()
    print("  Captured 01-dashboard.png")


# ══════════════════════════════════════════════════════════════════════
# Screenshot 2 — Eligible List (top rows)
# ══════════════════════════════════════════════════════════════════════
def capture_eligible():
    wb = load_workbook(WB_PATH)
    ws = wb["My Eligible Scholarships"]

    headers = ["#", "Scholarship", "Organization", "Status", "Deadline",
               "Match", "Benefits (short)"]
    col_widths = [32, 320, 200, 80, 160, 60, 340]
    rows = []
    for r in range(2, min(ws.max_row + 1, 9)):  # top 7 eligible
        vals = [
            ws.cell(r, 1).value or "",
            ws.cell(r, 2).value or "",
            ws.cell(r, 3).value or "",
            ws.cell(r, 4).value or "",
            ws.cell(r, 6).value or "",
            ws.cell(r, 16).value or "",
            (ws.cell(r, 10).value or "")[:80],
        ]
        score = vals[5]
        try:
            bg = GREEN_BG if int(score) >= 7 else YELLOW_BG if int(score) >= 4 else None
        except:
            bg = None
        rows.append([
            (str(vals[0]), LIGHT_BG, NAVY, FONT_BOLD_SM),
            (str(vals[1]), None, DARK_TEXT, FONT_SM),
            (str(vals[2]), None, DARK_TEXT, FONT_SM),
            (str(vals[3]), GREEN_BG if vals[3] == "Open" else None,
             GREEN_FG if vals[3] == "Open" else DARK_TEXT, FONT_BOLD_SM),
            (str(vals[4]), None, DARK_TEXT, FONT_SM),
            (str(vals[5]), bg, NAVY, FONT_BOLD),
            (str(vals[6]), None, DARK_TEXT, FONT_SM),
        ])

    img = render_table("Eligible Scholarships (Top Priority)", headers, rows, col_widths)
    img.save(os.path.join(OUT_DIR, "02-eligible-list.png"))
    wb.close()
    print("  Captured 02-eligible-list.png")


# ══════════════════════════════════════════════════════════════════════
# Screenshot 3 — Application Tracker
# ══════════════════════════════════════════════════════════════════════
def capture_tracker():
    wb = load_workbook(WB_PATH)
    ws = wb["Application Tracker"]

    headers = ["Scholarship", "Status", "Deadline", "Priority"]
    col_widths = [380, 160, 180, 80]
    rows = []
    max_r = min(ws.max_row + 1, 11)  # header + 10 rows
    for r in range(2, max_r):
        vals = [
            ws.cell(r, 1).value or "",
            ws.cell(r, 2).value or "",
            ws.cell(r, 3).value or "",
            ws.cell(r, 7).value or "",
        ]
        # Status color
        st = str(vals[1])
        st_bg = None
        st_fg = DARK_TEXT
        if st == "Awarded":
            st_bg = GREEN_BG
            st_fg = GREEN_FG
        elif st == "Preparing":
            st_bg = YELLOW_BG
            st_fg = YELLOW_FG
        elif "Monitor" in st:
            st_bg = (242, 242, 242)
            st_fg = GRAY_TEXT

        pri = str(vals[3])
        pri_fg = {"High": (192, 0, 0), "Medium": (191, 143, 0), "Low": GRAY_TEXT}.get(pri, DARK_TEXT)

        rows.append([
            (str(vals[0]), None, DARK_TEXT, FONT_SM),
            (str(vals[1]), st_bg, st_fg, FONT_BOLD_SM),
            (str(vals[2]), None, DARK_TEXT, FONT_SM),
            (str(vals[3]), None, pri_fg, FONT_BOLD_SM),
        ])

    img = render_table("Application Tracker", headers, rows, col_widths)
    img.save(os.path.join(OUT_DIR, "03-application-tracker.png"))
    wb.close()
    print("  Captured 03-application-tracker.png")


# ══════════════════════════════════════════════════════════════════════
# Screenshot 4 — Ranked Prioritization
# ══════════════════════════════════════════════════════════════════════
def capture_ranked():
    wb = load_workbook(WB_PATH)
    ws = wb["Ranked Prioritization"]

    headers = ["Rank", "Scholarship", "Organization", "Status", "Score", "Your Eligibility"]
    col_widths = [48, 340, 200, 80, 60, 120]
    rows = []
    for r in range(2, min(ws.max_row + 1, 12)):
        vals = [
            ws.cell(r, 1).value or "",
            ws.cell(r, 2).value or "",
            ws.cell(r, 3).value or "",
            ws.cell(r, 4).value or "",
            ws.cell(r, 6).value or "",
            ws.cell(r, 8).value or "",
        ]
        elig = str(vals[5])
        if elig == "Eligible":
            e_bg = GREEN_BG
            e_fg = GREEN_FG
        elif elig == "Conditional":
            e_bg = YELLOW_BG
            e_fg = YELLOW_FG
        else:
            e_bg = None
            e_fg = DARK_TEXT

        score = str(vals[4])
        try:
            s_bg = GREEN_BG if int(score) >= 7 else YELLOW_BG if int(score) >= 4 else None
        except:
            s_bg = None

        rows.append([
            (score, s_bg, NAVY, FONT_BOLD),
            (str(vals[1]), None, DARK_TEXT, FONT_SM),
            (str(vals[2]), None, DARK_TEXT, FONT_SM),
            (str(vals[3]), None, DARK_TEXT, FONT_SM),
            (score, s_bg, NAVY, FONT_BOLD),
            (str(vals[5]), e_bg, e_fg, FONT_BOLD_SM),
        ])

    img = render_table("Ranked Prioritization", headers, rows, col_widths)
    img.save(os.path.join(OUT_DIR, "04-ranked-priorities.png"))
    wb.close()
    print("  Captured 04-ranked-priorities.png")


if __name__ == "__main__":
    print("Capturing screenshots...")
    capture_dashboard()
    capture_eligible()
    capture_tracker()
    capture_ranked()
    print(f"\nAll screenshots saved to {OUT_DIR}")
