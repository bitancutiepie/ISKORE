"""
Build iskolar-tracker.xlsx — Scholarship Portfolio Tracker
Refactored with configurable USER PROFILE block.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── USER PROFILE ── Edit these for your own use
PROFILE = {
    "name": "Demo Scholar",
    "course": "BSIT",
    "year": "Incoming 3rd Year",
    "school": "Batangas State University",
    "municipality": "Batangas Province",
    "is_female": False,
    "held_awards": ["LGU Educational Grant", "Provincial Merit Scholarship"],
}

# ── Eligibility Engine ──
def determine_eligibility(s, profile):
    """
    Determine your eligibility for a scholarship based on your profile.
    Returns "Eligible", "Ineligible", or "Conditional".
    """
    course_ok = s["bsit"] == "Yes" and profile["course"] == "BSIT"
    yr_ok = s["yr3"] == "Yes" and "3rd" in profile["year"]
    region = s["region"]
    mun = profile["municipality"]
    is_female = profile["is_female"]
    held = profile["held_awards"]

    if s["bsit"] == "No (Statistics only)":
        return "Ineligible"
    if s["bsit"] == "No (Medicine only)":
        return "Ineligible"
    if s["bsit"] == "No (Nursing/Allied Health only)":
        return "Ineligible"
    if s["bsit"] == "No (Electromechanics)":
        return "Ineligible"
    if s["bsit"] == "No (specific Eng'g/Accountancy)":
        return "Ineligible"
    if s["bsit"] == "No (CS, ECE, CpE at Ateneo)":
        return "Ineligible"
    if s["bsit"] == "No (Statistics/OR)":
        return "Ineligible"
    if s["bsit"] == "No (Education majors only)":
        return "Ineligible"
    if s["bsit"] == "No (Architecture only)":
        return "Ineligible"

    if not course_ok:
        return "Ineligible"

    if s["yr3"] in ("No (freshman only)", "No (incoming sophomores)",
                     "No (2-yr tech-voc)", "No (freshmen abroad)",
                     "No (mobility program)", "No (incoming 2nd yr)",
                     "No (Grade 12 only)", "No (graduating seniors)",
                     "No (MD students)", "No (freshman entry)"):
        return "Ineligible"
    if not yr_ok and s["yr3"] == "No (freshman only)":
        return "Ineligible"

    if is_female and "female" in s["notes"].lower() and "ineligible" in s["notes"].lower():
        return "Ineligible"

    region_lower = region.lower()
    mun_lower = mun.lower()
    if "batangas city" in region_lower and "batangas city" not in mun_lower:
        return "Ineligible"
    if "san juan" in region_lower and "san juan" not in mun_lower:
        return "Ineligible"
    if "must move to" in region_lower:
        return "Ineligible"
    if "partner schools" in region_lower or "partner" in region_lower:
        schools = region_lower.replace("partner schools only (", "").replace(")", "").replace("partner schools (", "").replace(")", "")
        if profile["school"].lower() not in schools:
            return "Ineligible"

    if "no double-dipping" in s["notes"].lower() or "no other scholarship" in s["notes"].lower() or "prohibits holding" in s["notes"].lower():
        if held:
            return "Ineligible"

    if s["name"] == "DBP INSPIRE Scholarship":
        return "Conditional"
    if "coconut farmer" in s["notes"].lower():
        return "Conditional"
    if "gawad" in s["notes"].lower():
        return "Conditional"

    return "Eligible"

# ── colour palette & styles ──────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
BODY_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", size=10, bold=True)
LINK_FONT     = Font(name="Arial", size=10, color="0563C1", underline="single")
ELIGIBLE_FILL = PatternFill("solid", fgColor="C6EFCE")
INELIG_FILL   = PatternFill("solid", fgColor="FFC7CE")
COND_FILL     = PatternFill("solid", fgColor="FFEB9C")
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER   = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_body(ws, max_row, max_col, alt=True):
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if alt and r % 2 == 0:
                if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = ALT_ROW_FILL

def auto_width(ws, min_w=8, max_w=45):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(min(len(str(cell.value)), max_w))
        best = max(lengths) + 3 if lengths else min_w
        ws.column_dimensions[col_letter].width = max(min_w, min(best, max_w))

def style_elig_cell(cell, value):
    if value == "Eligible":
        cell.fill = ELIGIBLE_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="006100")
    elif value == "Ineligible":
        cell.fill = INELIG_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
    elif value == "Conditional":
        cell.fill = COND_FILL
        cell.font = Font(name="Arial", size=10, bold=True, color="9C6500")

def write_link(cell, url):
    if url and str(url).startswith("http"):
        cell.font = LINK_FONT
        cell.hyperlink = str(url)

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────
# SCHOLARSHIP DATA — all 47 entries
# Fields: name, org, status, open_date, deadline, yr3, bsit, region,
#         benefits, income, gpa, exam, interview, competitiveness,
#         match_score, elap_ok, prov_ok, link, notes, your_eligible
# ─────────────────────────────────────────────────────────────────────────

S = []

def add(name, org, status, open_date, deadline, yr3, bsit, region,
        benefits, income, gpa, exam, interview, comp, match_score,
        elap_ok, prov_ok, link, notes, your_eligible):
    S.append(dict(
        name=name, org=org, status=status, open_date=open_date,
        deadline=deadline, yr3=yr3, bsit=bsit, region=region,
        benefits=benefits, income=income, gpa=gpa, exam=exam,
        interview=interview, comp=comp, match_score=match_score,
        elap_ok=elap_ok, prov_ok=prov_ok, link=link, notes=notes,
        your_eligible=your_eligible,
    ))

# --- 1-3: DOST JLSS tracks ---
add("DOST JLSS — RA 7687 Track", "DOST-SEI", "Closed", "Apr 13, 2026", "May 15, 2026",
    "Yes", "Yes", "None (National)", "₱8,000/mo living allowance, ₱10k/yr books, ₱10k thesis, health insurance, ₱1k graduation, travel allowance",
    "Disadvantaged families", "GWA ≥ 83%", "Yes", "No", "Very High", 10,
    "Not stated", "Not stated",
    "https://jlss.science-scholarships.ph",
    "Top priority. Requires relinquishing other nat'l merit grants if awarded. Exam covers advanced physics/chem/calculus.",
    "Eligible")

add("DOST JLSS — Merit Track", "DOST-SEI", "Closed", "Apr 13, 2026", "May 15, 2026",
    "Yes", "Yes", "None (National)", "₱8,000/mo living allowance, ₱10k/yr books, ₱10k thesis, health insurance",
    "None (merit-based)", "GWA ≥ 83%", "Yes", "No", "Very High", 10,
    "Not stated", "Not stated",
    "https://jlss.science-scholarships.ph",
    "Perfect for strong GWA. No income barrier. Same exam as RA 7687.",
    "Eligible")

add("DOST JLSS — RA 10612 Track", "DOST-SEI", "Closed", "Apr 13, 2026", "May 15, 2026",
    "Yes", "Yes", "None (National)", "₱8,000/mo living allowance, ₱10k/yr books, guaranteed teaching job placement",
    "None", "GWA ≥ 83%", "Yes", "No", "High", 9,
    "Not stated", "Not stated",
    "https://jlss.science-scholarships.ph",
    "Mandatory teaching service may divert from SWE goals. Slightly lower match.",
    "Eligible")

# --- 4: GBF STEM ---
add("GBF STEM-College Scholarship", "Gokongwei Brothers Foundation", "Closed", "Mar/Apr 2026", "May 31, 2026",
    "Yes", "Yes", "None (National)", "Up to ₱120k/yr, leadership training, internship/employment at Gokongwei Group",
    "Financial need", "GWA ≥ 85% / 2.0", "Yes", "Yes", "High", 9,
    "Yes (case-by-case)", "Yes",
    "https://bit.ly/GBFSTEMCollege",
    "Strong fit for SWE goals. Return service at Gokongwei Group. Portfolio matters.",
    "Eligible")

# --- 5: Presidential BBM ---
add("Presidential Scholars Program (B BBM)", "Office of the President / Batangas LGU", "Open", "Apr 24, 2026", "Rolling (barangay allocation)",
    "Yes", "Yes", "Batangas Province (5 per barangay)", "₱20,000 educational assistance",
    "Low-income priority", "Passing standing (not explicitly stated)", "No", "No", "Medium", 9,
    "Yes", "Yes",
    "Coordinated at Barangay Hall",
    "Active rolling. Visit your Barangay Captain. Submit early.",
    "Eligible")

# --- 6: Batangas City EBD ---
add("Batangas City EBD Scholarship", "Batangas City Government", "Open", "Jul 2026", "Aug 2026",
    "Yes", "Yes", "Batangas City resident ≥5 yrs", "₱3,000/sem allowance (SUC scholars)",
    "Marginally poor", "GWA ≥ 2.5 (80%)", "No", "Yes (phone)", "Medium-High", 9,
    "Not stated (generally allowed)", "Not stated",
    "ebd_scholarship@batangascity.gov.ph",
    "⚠ INELIGIBLE — resident must be Batangas City.",
    "Ineligible")

# --- 7: CHED TDP ---
add("CHED Tulong Dunong (TDP)", "CHED", "Open", "Jul 13, 2026", "Mid-late Jul 2026 (3-4 day window)",
    "Yes", "Yes", "Legislative district of sponsoring rep", "₱15,000/yr",
    "Marginally disadvantaged", "Passing average", "No", "No", "High", 9,
    "Not stated", "Not stated",
    "https://ched.gov.ph",
    "Facilitated through congressional district office. Very short submission window.",
    "Eligible")

# --- 8: CHED TES ---
add("CHED Tertiary Education Subsidy (TES)", "CHED / UniFAST", "Open", "Jun/Jul 2026", "Determined by SUC enrollment",
    "Yes", "Yes", "None (enrolled in SUC)", "₱20,000–₱40,000/yr",
    "DSWD Listahanan / 4Ps priority", "Satisfactory standing", "No", "No", "Extremely High", 9,
    "Not stated", "Not stated",
    "https://unifast.gov.ph",
    "Internal via university scholarship office. Non-Listahanan = lower priority.",
    "Eligible")

# --- 9: Fuji Haya ---
add("Fuji Haya Electric Scholarship", "Fuji Haya Electric / BatStateU", "Open", "Enrollment cycle", "Semestral deadlines",
    "Yes", "Yes", "BatStateU student", "Semestral grant, seminars, OJT, potential employment",
    "Financial need", "High standing", "No", "Yes", "Medium-High", 8,
    "Not stated", "Not stated",
    "https://global.batstateu.edu.ph",
    "Directly at campus. Prohibits other private corporate awards but LGU ok.",
    "Eligible")

# --- 10: Generation Google ---
add("Generation Google (APAC) — Women in CS", "Google Inc.", "Open", "Jul 1, 2026", "Aug 14, 2026",
    "Yes", "Yes", "None (APAC region)", "$2,500 USD (~₱140,000) pure stipend",
    "Financial need considered", "GWA ≥ 85%", "Yes (Online Challenge)", "Yes", "Extremely High", 8,
    "Yes", "Yes",
    "https://buildyourfuture.withgoogle.com",
    "⚠ INELIGIBLE — female-identifying applicants only.",
    "Ineligible")

# --- 11: Megaworld ---
add("Megaworld Foundation Scholarship", "Megaworld Foundation", "Closed", "Feb 11, 2026", "Jul 31, 2026",
    "Yes", "Yes", "Partner schools (BatStateU is partner)", "Full tuition waiver, monthly living allowance, seminars, career placement",
    "Income ≤ ₱400k/yr", "GWA ≥ 85%, no subject <80%", "Yes", "Yes", "High", 8,
    "No", "No",
    "https://www.megaworldfoundation.com/scholarship/apply/form",
    "Strict 'no other scholarship' clause conflicts with held awards.",
    "Ineligible")

# --- 12: DBP INSPIRE ---
add("DBP INSPIRE Scholarship", "Development Bank of the Philippines", "Open", "Mid-year 2026", "Not explicitly stated",
    "Yes", "Yes", "None (PH-wide SUC network)", "Semestral allowance + learning materials subsidy",
    "Marginally disadvantaged", "Passing standing", "No", "No", "High", 4,
    "Not stated", "Not stated",
    "https://www.dbp.ph",
    "Limited info on SUC quotas. Requires verification with university scholarship office.",
    "Conditional")

# --- 13: DICT STEM/ICT ---
add("DICT STEM/ICT Scholarship", "DICT", "Open", "Rolling", "Aligns with enrollment",
    "Yes", "Yes", "None", "Specialized training, global certifications, upskilling",
    "Disadvantaged families", "Passing marks", "No", "No", "Medium-High", 7,
    "Not stated", "Not stated",
    "https://dict.gov.ph",
    "High alignment with SWE goals. Certification-focused rather than cash stipend.",
    "Eligible")

# --- 14: Accenture Internship ---
add("Accenture Internship & Sponsorship", "Accenture Philippines", "Open", "Rolling", "Rolling",
    "Yes", "Yes", "None", "Upskilling (ABAP, SAP, Salesforce), internship stipend, priority hiring",
    "None", "Satisfactory progress", "Yes (Cognitive)", "Yes", "Medium", 7,
    "Yes", "Yes",
    "https://www.accenture.com/ph-en/careers/form/phstudents",
    "Career-development program, not traditional cash stipend. Excellent for SWE pipeline.",
    "Eligible")

# --- 15: San Juan Iskolar ---
add("San Juan Iskolar ng Bayan", "Municipal Gov of San Juan, Batangas", "Open", "Jul/Aug 2026", "Municipal registration deadline",
    "Yes", "Yes", "San Juan, Batangas resident", "Load assistance + semestral financial support",
    "Financially needy", "Passing grades", "No", "No", "Low-Medium", 6,
    "Yes", "Yes",
    "San Juan LGU Mayor's Office",
    "⚠ INELIGIBLE — resident must be San Juan.",
    "Ineligible")

# --- 16: CHED CMSP ---
add("CHED Merit Scholarship (CMSP)", "CHED", "Closed", "Jun 22, 2026", "Jul 31, 2026",
    "No (freshman only)", "Yes", "None", "₱40,000–₱80,000/yr",
    "Income ≤ ₱500k", "Grade 12 GWA ≥ 93%", "No", "No", "Extremely High", 1,
    "No", "No",
    "https://ched.gov.ph",
    "INELIGIBLE — restricted to incoming freshmen.",
    "Ineligible")

# --- 17: CHED BPMSP ---
add("CHED Bagong Pilipinas Merit (BPMSP)", "CHED", "Closed", "Jun 2026", "Jul 2026",
    "No (freshman only)", "Yes", "None", "Full tuition support (private HEIs), stipend, living allowance",
    "Income ≤ ₱2M", "Grade 12 GWA ≥ 95%", "No", "No", "Extremely High", 1,
    "No", "No",
    "https://bpms.ched.gov.ph",
    "INELIGIBLE — restricted to first-time college entrants.",
    "Ineligible")

# --- 18: CHED Estatistikolar ---
add("CHED Estatistikolar", "CHED", "Closed", "Jun 22, 2026", "Aug 15, 2026",
    "Yes", "No (Statistics only)", "None", "₱35,000/sem stipend + ₱5k books",
    "Income ≤ ₱500k", "Passing", "No", "No", "High", 4,
    "Not stated", "Not stated",
    "https://ched.gov.ph",
    "INELIGIBLE — restricted to BS Statistics / Applied Statistics.",
    "Ineligible")

# --- 19: CHED CoScho ---
add("CHED Coconut Farmers (CoScho)", "CHED", "Open", "Jun 22, 2026", "Jul 31, 2026",
    "Yes", "Yes", "None", "Up to ₱40,000/sem allowance",
    "Registered coconut farmer/dependent", "Satisfactory standing", "No", "No", "High", 3,
    "Not stated", "Not stated",
    "https://ched.gov.ph",
    "Requires NCFRS registration. Conditional on parent being a registered coconut farmer.",
    "Conditional")

# --- 20: CHED MSRS ---
add("CHED Medical Scholarship (MSRS)", "CHED", "Open", "Jun/Jul 2026", "Not stated",
    "No (MD students)", "No (Medicine only)", "Partner HEIs", "Full tuition, living/book/transport allowance",
    "Indigent", "NMAT + med school standards", "Yes (NMAT)", "Yes", "Extremely High", 1,
    "No", "No",
    "https://ched.gov.ph",
    "INELIGIBLE — Doctor of Medicine program only.",
    "Ineligible")

# --- 21: CHED AHEAD ---
add("CHED AHEAD Grant", "CHED", "Open", "Jul 2026", "School-specific",
    "Yes", "No (Nursing/Allied Health only)", "None", "One-time ₱25,000 clinical training assistance",
    "Financial need", "Satisfactory standing", "No", "No", "Medium", 1,
    "No", "No",
    "https://ched.gov.ph",
    "INELIGIBLE — restricted to Nursing and Allied Health students.",
    "Ineligible")

# --- 22: CHED SIDA-SGP ---
add("CHED SIDA-SGP", "CHED", "Open", "Jun 2026", "Regional schedules",
    "No (freshman entry)", "Yes", "Special development areas", "Semestral cash subsidy, tuition, book support",
    "Low-income", "Passing marks", "No", "No", "High", 1,
    "No", "No",
    "https://ched.gov.ph",
    "INELIGIBLE — restricted to specific regions and incoming freshmen.",
    "Ineligible")

# --- 23: Aboitiz Future Leaders ---
add("Aboitiz Future Leaders Scholarship", "Aboitiz Foundation", "Closed", "Aug 1, 2025/26", "Sep 1, 2025/26",
    "No (incoming sophomores)", "Yes", "Partner universities", "Full tuition, monthly stipend, 400-hr internship",
    "Financial need", "GWA ≥ 88%", "Yes", "Yes", "High", 1,
    "No", "No",
    "https://aboitiz.com",
    "INELIGIBLE — open only to incoming 2nd year students.",
    "Ineligible")

# --- 24: Aboitiz Dualtech ---
add("Aboitiz Dualtech Scholarship", "Aboitiz / Dualtech", "Open", "Rolling", "Rolling",
    "No (2-yr tech-voc)", "No (Electromechanics)", "None", "Subsidized training, OJT, tools",
    "Indigent young men", "Basic passing", "Yes", "Yes", "Medium", 1,
    "No", "No",
    "https://www.dualtech.org.ph",
    "INELIGIBLE — 2-year tech-voc program, not BSIT.",
    "Ineligible")

# --- 25: Lao Dualtech ---
add("Lao Foundation Dualtech Scholarship", "Lao Foundation / Dualtech", "Open", "Rolling", "Rolling",
    "No (2-yr tech-voc)", "No (Electromechanics)", "None", "Subsidized training fees, industrial placement",
    "Indigent youth", "Basic passing", "Yes", "Yes", "Medium", 1,
    "No", "No",
    "https://www.dualtech.org.ph",
    "INELIGIBLE — 2-year tech-voc program, not BSIT.",
    "Ineligible")

# --- 26: SM Foundation ---
add("SM Foundation College Scholarship", "SM Foundation", "Closed", "Dec-Feb", "Not stated",
    "No (Grade 12 only)", "Yes", "Partner schools", "Full tuition, monthly allowance, part-time jobs",
    "Income ≤ ₱250k", "Grade 12 GWA ≥ 92%", "Yes", "Yes", "Extremely High", 1,
    "No", "No",
    "https://www.sm-foundation.org",
    "INELIGIBLE — restricted to graduating Grade 12 students.",
    "Ineligible")

# --- 27: Security Bank External ---
add("Security Bank External College Scholarship", "Security Bank Foundation", "Closed", "May 28, 2025/26", "Jun 30, 2025/26",
    "Yes (limited slots)", "Yes", "8 partner schools only (ADMU, DLSU, FEU, PLM, PUP, UPD, UST, CKSC)", "₱100,000/yr stipend (SUC scholars)",
    "Underprivileged", "GWA ≥ 93%, no subject <86%", "No", "No (validation call)", "Very High", 1,
    "No", "No",
    "https://www.securitybank.com/foundation",
    "INELIGIBLE — BatStateU is not a partner school.",
    "Ineligible")

# --- 28: Security Bank RMKK ---
add("Security Bank RMKK Agency Personnel Scholarship", "Security Bank Foundation", "Open", "Not stated", "Not stated",
    "Yes", "Yes", "None", "Up to ₱30,000/yr",
    "SB agency personnel only", "GWA ≥ 83%", "No", "No", "Medium", 1,
    "No", "No",
    "https://www.securitybank.com",
    "INELIGIBLE — restricted to Security Bank agency employees.",
    "Ineligible")

# --- 29: Security Bank Employee Dependent ---
add("Security Bank Employee Dependent Internal Scholarship", "Security Bank Foundation", "Open", "Jun 2026", "Jun 30, 2026",
    "Yes", "Yes", "None", "Direct tuition and stipend assistance",
    "SB permanent employees ≥2 yrs", "GWA ≥ 83%, min 83% per subject", "No", "No", "Medium", 1,
    "No", "No",
    "https://www.securitybank.com",
    "INELIGIBLE — restricted to children of SB employees.",
    "Ineligible")

# --- 30: BPI Pagpupugay ---
add("BPI Foundation Pagpupugay Scholarship", "BPI Foundation", "Closed", "Jun 2026", "Closed",
    "Yes", "Yes", "None", "Up to ₱100,000/yr",
    "Next of kin of COVID frontliner", "GWA ≥ 85%", "No", "No", "High", 2,
    "Not stated", "Not stated",
    "https://www.bpifoundation.org",
    "INELIGIBLE — requires being next of kin of a medical frontliner.",
    "Ineligible")

# --- 31: BPI-DOST Innovation ---
add("BPI-DOST Innovation Awards", "BPI Foundation & DOST-STII", "Closed", "Jan 2026", "Jun 30, 2026",
    "Yes", "Yes", "None", "₱60k–₱150k cash prizes (team competition)",
    "None", "Regular enrollment, no failing grades", "No", "Yes (presentation)", "Extremely High", 7,
    "Yes", "Yes",
    "https://www.bpifoundation.org",
    "Group project competition. Requires dean endorsement and 3-member team.",
    "Eligible")

# --- 32: InLife Gold Eagle ---
add("InLife Gold Eagle College Scholarship", "Insular Life Foundation", "Closed", "Feb 6, 2026", "Jun 30, 2026",
    "No (freshman only)", "Yes", "Partner SUCs (UPD, ASCOT, BSU, BISU, CMU, DORSU)", "Semestral stipends, book/transport allowances",
    "Income ≤ ₱250k (provincial SUCs)", "Grade 11-12 GWA ≥ 85%", "No", "No", "High", 1,
    "No", "No",
    "https://www.insularfoundation.com.ph",
    "INELIGIBLE — restricted to incoming 1st year college students.",
    "Ineligible")

# --- 33: PHINMA PNS ---
add("PHINMA National Scholarship (PNS)", "PHINMA Foundation", "Closed", "Jun 12, 2026", "Jul 15, 2026",
    "No (incoming 2nd yr)", "No (specific Eng'g/Accountancy)", "5 partner schools (UPD, PUP, PNU, TUP, UPang)", "Monthly stipend, leadership training, mentorship",
    "Income ≤ ₱300k", "GWA ≥ 80%", "No", "Yes", "Very High", 1,
    "No", "No",
    "https://www.phinmafoundation.org",
    "INELIGIBLE — wrong year level, wrong course, wrong SUC.",
    "Ineligible")

# --- 34: Huawei Seeds ---
add("Huawei Seeds for the Future", "Huawei Technologies", "Open", "Not stated", "Program-specific",
    "Yes", "Yes", "None", "All-expenses-paid tech training, global competition, cert vouchers",
    "None (merit-based)", "Strong academic records", "No", "Yes", "High", 5,
    "Yes", "Yes",
    "https://www.huawei.com/minisite/seeds-for-the-future",
    "Short-term training program, not ongoing cash stipend. Great for upskilling.",
    "Eligible")

# --- 35: Huawei-DLSU ---
add("Huawei-DLSU Engineering & Tech Scholarship", "Huawei / DLSU", "Closed", "Academic year start", "Semestral cycle",
    "Yes", "Yes", "DLSU Manila student", "Semestral tuition + cash allowance",
    "Financial need + merit", "DLSU standards", "No", "Yes", "High", 1,
    "No", "No",
    "https://www.dlsu.edu.ph",
    "INELIGIBLE — restricted to DLSU students.",
    "Ineligible")

# --- 36: Huawei-Ateneo ---
add("Huawei-Ateneo Science & Engineering Scholarship", "Huawei / Ateneo", "Open", "Enrollment aligned", "Ateneo schedule",
    "Yes", "No (CS, ECE, CpE at Ateneo)", "Ateneo de Manila student", "Tuition waiver + monthly living support",
    "Merit + financial need", "High average", "No", "Yes", "High", 1,
    "No", "No",
    "https://www.ateneo.edu",
    "INELIGIBLE — restricted to Ateneo students in specific programs.",
    "Ineligible")

# --- 37: Accenture Analytics ---
add("Accenture Analytics Career Path Scholarship", "Accenture / UPC", "Closed", "Not stated", "Closed",
    "Yes", "No (Statistics/OR)", "UPC Barcelona, Spain", "Enrollment credits + admin fees",
    "None", "High standing", "No", "No", "Very High", 1,
    "No", "No",
    "https://www.wemakescholars.com",
    "INELIGIBLE — restricted to UPC Barcelona and Statistics/OR majors.",
    "Ineligible")

# --- 38: Study in Saudi Arabia ---
add("Study in Saudi Arabia Program", "Saudi Ministry of Education", "Open", "International schedule", "University-dependent",
    "No (freshmen abroad)", "Yes", "Must move to Saudi Arabia", "Full tuition waiver, living allowance, housing, medical",
    "None", "Excellent standing", "No", "No", "High", 1,
    "No", "No",
    "https://studyinsaudi.moe.gov.sa",
    "INELIGIBLE — requires relocating abroad as a freshman.",
    "Ineligible")

# --- 39: AMEXCID Mexico ---
add("AMEXCID Merit Scholarship", "Mexican Agency for International Cooperation", "Open", "Academic cycle", "Varies",
    "No (mobility program)", "Yes", "Must move to Mexico", "Living allowance, tuition, health insurance, travel",
    "None", "GWA ≥ 85%", "No", "No", "Very High", 1,
    "No", "No",
    "https://www.gob.mx/amexcid",
    "INELIGIBLE — requires relocating to Mexico.",
    "Ineligible")

# --- 40: CSC Gawad PASUC ---
add("Gawad Lingkod Bayani PASUC", "CSC & PASUC", "Open", "May 25, 2026", "Not stated",
    "Yes", "Yes", "PASUC-member SUCs (incl. BatStateU)", "Educational subsidies at any PASUC SUC",
    "Not stated", "Not stated", "No", "No", "Low-Medium", 5,
    "Yes", "Yes",
    "https://www.csc.gov.ph",
    "Restricted to Gawad awardees and their immediate kin. Conditional on family status.",
    "Conditional")

# --- 41: Metrobank ACCESS ---
add("Metrobank ACCESS Scholarship", "Metrobank Foundation", "Closed", "Enrollment cycle", "Partner timeline",
    "No (freshman only)", "Yes", "Partner PAASCU Level II/III schools", "Tuition + fixed semestral allowances + workshops",
    "Income ≤ ₱500k", "GWA ≥ 85%", "Yes", "Yes", "Extremely High", 1,
    "No", "No",
    "https://www.mbfoundation.org.ph",
    "INELIGIBLE — restricted to incoming freshmen.",
    "Ineligible")

# --- 42: Metrobank ETP ---
add("Metrobank Excellence in Teaching (ETP)", "Metrobank Foundation", "Closed", "Academic year start", "Enrollment season",
    "Yes", "No (Education majors only)", "COE SUCs", "Stipends, tuition support, development programs",
    "Poorest regions", "GWA ≥ 85%", "No", "Yes", "High", 1,
    "No", "No",
    "https://www.mbfoundation.org.ph",
    "INELIGIBLE — restricted to Education majors.",
    "Ineligible")

# --- 43: Metrobank GRACE ---
add("Metrobank GRACE Grant", "Metrobank Foundation", "Closed", "Final year start", "School-specific",
    "No (graduating seniors)", "Yes", "Partner Level II/III schools", "Semestral cash subsidy",
    "High drop-out risk", "GWA ≥ 85%", "No", "Yes", "High", 1,
    "No", "No",
    "https://www.mbfoundation.org.ph",
    "INELIGIBLE — restricted to graduating seniors.",
    "Ineligible")

# --- 44: Metrobank MSP ---
add("Metrobank Scholarship Program (MSP)", "Metrobank Foundation", "Closed", "Academic year start", "Enrollment season",
    "No (freshman only)", "Yes", "Partner schools", "Semestral cash stipends + tuition waivers",
    "Marginalized households", "Satisfactory progress", "Yes", "Yes", "Very High", 1,
    "No", "No",
    "https://www.mbfoundation.org.ph",
    "INELIGIBLE — restricted to incoming freshmen.",
    "Ineligible")

# --- 45: Metrobank Boysen ---
add("Metrobank-Boysen Paint Scholarship", "Metrobank Foundation & Boysen", "Open", "Enrollment aligned", "Not stated",
    "Yes", "No (Architecture only)", "Architecture partner schools", "Semestral cash stipends, design project support, books",
    "Underprivileged", "GWA ≥ 85%", "No", "Yes", "Very High", 1,
    "No", "No",
    "https://www.mbfoundation.org.ph",
    "INELIGIBLE — restricted to Architecture majors.",
    "Ineligible")

# --- 46: Ayala U-Go ---
add("Ayala Foundation U-Go Scholar Grant", "Ayala Foundation / U-Go Global", "Closed", "May 5, 2026", "Jun 6, 2026",
    "Yes", "Yes", "Public/state universities", "₱40,000/yr + leadership training + professional development",
    "Low-income", "GWA ≥ 85%", "No", "Yes", "Very High", 1,
    "No", "No",
    "https://ayalafoundation.org/programs/scholarships",
    "INELIGIBLE — female-prioritized AND prohibits holding other active grants.",
    "Ineligible")

# --- 47: GSIS GSSP ---
add("GSIS Subsidy for STEM (GSSP)", "GSIS", "Closed", "Jun 30, 2026", "Aug 31, 2026",
    "No (freshman only)", "Yes", "None", "₱15,000/yr + Latin honors incentives up to ₱50k",
    "Parent GSIS member SG ≤15", "Grade 12 GWA ≥ 90%", "No", "No", "High", 1,
    "No", "No",
    "https://www.gsis.gov.ph",
    "INELIGIBLE — restricted to incoming first-time college freshmen.",
    "Ineligible")


# ═════════════════════════════════════════════════════════════════════════
# COMPUTED STATS
# ═════════════════════════════════════════════════════════════════════════
total = len(S)
open_count = sum(1 for s in S if s["status"] == "Open")
eligible = sum(1 for s in S if s["your_eligible"] == "Eligible")
conditional = sum(1 for s in S if s["your_eligible"] == "Conditional")
ineligible = sum(1 for s in S if s["your_eligible"] == "Ineligible")
high_match = sum(1 for s in S if s["match_score"] >= 7 and s["your_eligible"] == "Eligible")


# ═════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ═════════════════════════════════════════════════════════════════════════
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = "1F4E79"

ws_dash.merge_cells("A1:H1")
ws_dash["A1"] = "ISKOLAR TRACKER — Scholarship Portfolio"
ws_dash["A1"].font = Font(name="Arial", bold=True, size=18, color="1F4E79")
ws_dash["A1"].alignment = Alignment(horizontal="left")

ws_dash.merge_cells("A2:H2")
ws_dash["A2"] = f"{PROFILE['course']} @ {PROFILE['school']} · {PROFILE['year']} · {PROFILE['municipality']}"
ws_dash["A2"].font = Font(name="Arial", size=11, color="555555")
ws_dash["A2"].alignment = Alignment(horizontal="left")

ws_dash.merge_cells("A3:H3")
ws_dash["A3"] = f"Generated: {datetime.date.today().strftime('%B %d, %Y')}  |  Profile: {PROFILE['name']}"
ws_dash["A3"].font = Font(name="Arial", size=10, italic=True, color="999999")

# Currently held awards
ws_dash.merge_cells("A5:C5")
ws_dash["A5"] = "CURRENTLY HELD AWARDS"
ws_dash["A5"].font = SUBTITLE_FONT

held_awards_data = [
    (a, "Various", "Awarded") for a in PROFILE["held_awards"]
]
for i, (n, o, st) in enumerate(held_awards_data):
    r = 6 + i
    ws_dash[f"A{r}"] = n
    ws_dash[f"A{r}"].font = BOLD_FONT
    ws_dash[f"B{r}"] = o
    ws_dash[f"B{r}"].font = BODY_FONT
    ws_dash[f"C{r}"] = st
    ws_dash[f"C{r}"].font = Font(name="Arial", size=10, bold=True, color="006100")
    ws_dash[f"C{r}"].fill = PatternFill("solid", fgColor="C6EFCE")

# Key stats
stat_start = 6 + len(PROFILE["held_awards"]) + 1
ws_dash.merge_cells(f"A{stat_start}:C{stat_start}")
ws_dash[f"A{stat_start}"] = "PORTFOLIO SUMMARY"
ws_dash[f"A{stat_start}"].font = SUBTITLE_FONT

stats = [
    ("Total Scholarships Evaluated", total, None),
    ("Currently Open for Application", open_count, "E2EFDA"),
    ("Eligible for You (based on profile)", eligible, "C6EFCE"),
    ("Conditional Eligibility", conditional, "FFEB9C"),
    ("Ineligible", ineligible, "FFC7CE"),
    ("Top Priority Targets (Score ≥ 7 + Eligible)", high_match, None),
    ("Currently Holding (Active)", len(PROFILE["held_awards"]), None),
]
for i, (label, val, color) in enumerate(stats):
    r = stat_start + 1 + i
    ws_dash[f"A{r}"] = label
    ws_dash[f"A{r}"].font = BODY_FONT
    ws_dash[f"B{r}"] = val
    ws_dash[f"B{r}"].font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    ws_dash[f"B{r}"].alignment = Alignment(horizontal="center")
    if color:
        ws_dash[f"A{r}"].fill = PatternFill("solid", fgColor=color)
        ws_dash[f"B{r}"].fill = PatternFill("solid", fgColor=color)

# Upcoming deadlines
deadline_start = stat_start + 1 + len(stats) + 1
ws_dash.merge_cells(f"A{deadline_start}:D{deadline_start}")
ws_dash[f"A{deadline_start}"] = "UPCOMING DEADLINES (Eligible / Conditional)"
ws_dash[f"A{deadline_start}"].font = SUBTITLE_FONT

hd = deadline_start + 1
ws_dash[f"A{hd}"] = "Scholarship"
ws_dash[f"B{hd}"] = "Deadline"
ws_dash[f"C{hd}"] = "Status"
ws_dash[f"D{hd}"] = "Priority"
for c in range(1, 5):
    cell = ws_dash.cell(row=hd, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

deadlines = []
for s in S:
    if s["your_eligible"] in ("Eligible", "Conditional") and s["status"] == "Open":
        priority = "HIGH" if s["match_score"] >= 7 else "MEDIUM" if s["match_score"] >= 4 else "LOW"
        deadlines.append((s["name"], s["deadline"], s["status"], priority))
deadlines.sort(key=lambda x: x[1] if x[1] != "Rolling" else "ZZZZ")

for i, (name, dl, st, pri) in enumerate(deadlines):
    r = hd + 1 + i
    ws_dash[f"A{r}"] = name
    ws_dash[f"A{r}"].font = BODY_FONT
    ws_dash[f"B{r}"] = dl
    ws_dash[f"B{r}"].font = BODY_FONT
    ws_dash[f"C{r}"] = st
    ws_dash[f"C{r}"].font = BODY_FONT
    ws_dash[f"D{r}"] = pri
    ws_dash[f"D{r}"].font = Font(name="Arial", size=10, bold=True,
                                  color="C00000" if pri == "HIGH" else "BF8F00" if pri == "MEDIUM" else "555555")
    ws_dash[f"D{r}"].alignment = Alignment(horizontal="center")

# Legend
leg_r = hd + 1 + len(deadlines) + 1
ws_dash.merge_cells(f"A{leg_r}:D{leg_r}")
ws_dash[f"A{leg_r}"] = "ELIGIBILITY LEGEND"
ws_dash[f"A{leg_r}"].font = SUBTITLE_FONT

for i, (label, color) in enumerate([
    ("Eligible — ready to apply", "C6EFCE"),
    ("Conditional — depends on specific criteria", "FFEB9C"),
    ("Ineligible — does not match your profile", "FFC7CE"),
]):
    r = leg_r + 1 + i
    ws_dash[f"A{r}"] = label
    ws_dash[f"A{r}"].font = BODY_FONT
    ws_dash[f"A{r}"].fill = PatternFill("solid", fgColor=color)

ws_dash.column_dimensions["A"].width = 50
ws_dash.column_dimensions["B"].width = 35
ws_dash.column_dimensions["C"].width = 20
ws_dash.column_dimensions["D"].width = 15
for c in "EFGH":
    ws_dash.column_dimensions[c].width = 12


# ═════════════════════════════════════════════════════════════════════════
# SHEET 2: MASTER LIST
# ═════════════════════════════════════════════════════════════════════════
ws_master = wb.create_sheet("Master List")
ws_master.sheet_properties.tabColor = "2E75B6"

headers_master = [
    "#", "Scholarship Name", "Organization", "Status", "Open Date", "Deadline",
    "Eligible 3rd Yr", "Eligible BSIT", "Region Restriction", "Benefits",
    "Income Requirement", "GPA Requirement", "Exam", "Interview",
    "Competitiveness", "Match Score", "Your Eligibility",
    "ELAP Compatible", "Provincial Compatible", "Application Link", "Notes",
]
for c, h in enumerate(headers_master, 1):
    ws_master.cell(row=1, column=c, value=h)
style_header(ws_master, len(headers_master))

for i, s in enumerate(S):
    r = i + 2
    vals = [
        i + 1, s["name"], s["org"], s["status"], s["open_date"], s["deadline"],
        s["yr3"], s["bsit"], s["region"], s["benefits"],
        s["income"], s["gpa"], s["exam"], s["interview"],
        s["comp"], s["match_score"], s["your_eligible"],
        s["elap_ok"], s["prov_ok"], s["link"], s["notes"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws_master.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 20:
            write_link(cell, v)
    style_elig_cell(ws_master.cell(row=r, column=17), s["your_eligible"])

ws_master.auto_filter.ref = f"A1:{get_column_letter(len(headers_master))}{len(S)+1}"
ws_master.freeze_panes = "A2"
auto_width(ws_master, max_w=50)


# ═════════════════════════════════════════════════════════════════════════
# SHEET 3: MY ELIGIBLE SCHOLARSHIPS
# ═════════════════════════════════════════════════════════════════════════
ws_eligible = wb.create_sheet("My Eligible Scholarships")
ws_eligible.sheet_properties.tabColor = "548235"

eligible_only = [s for s in S if s["your_eligible"] == "Eligible"]
eligible_only.sort(key=lambda s: -s["match_score"])

for c, h in enumerate(headers_master, 1):
    ws_eligible.cell(row=1, column=c, value=h)
style_header(ws_eligible, len(headers_master))

for i, s in enumerate(eligible_only):
    r = i + 2
    vals = [
        i + 1, s["name"], s["org"], s["status"], s["open_date"], s["deadline"],
        s["yr3"], s["bsit"], s["region"], s["benefits"],
        s["income"], s["gpa"], s["exam"], s["interview"],
        s["comp"], s["match_score"], s["your_eligible"],
        s["elap_ok"], s["prov_ok"], s["link"], s["notes"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws_eligible.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 20:
            write_link(cell, v)
    ws_eligible.cell(row=r, column=16).font = Font(name="Arial", size=10, bold=True, color="1F4E79")

ws_eligible.auto_filter.ref = f"A1:{get_column_letter(len(headers_master))}{len(eligible_only)+1}"
ws_eligible.freeze_panes = "A2"
auto_width(ws_eligible, max_w=50)


# ═════════════════════════════════════════════════════════════════════════
# SHEET 4: APPLICATION TRACKER
# ═════════════════════════════════════════════════════════════════════════
ws_tracker = wb.create_sheet("Application Tracker")
ws_tracker.sheet_properties.tabColor = "ED7D31"

track_headers = [
    "Scholarship", "Status", "Deadline", "Link",
    "Docs Needed", "Docs Ready?", "Priority", "Notes / Next Step",
]
for c, h in enumerate(track_headers, 1):
    ws_tracker.cell(row=1, column=c, value=h)
style_header(ws_tracker, len(track_headers))

track_items = [s for s in S if s["your_eligible"] in ("Eligible", "Conditional")]
track_items.sort(key=lambda s: (-s["match_score"], s["status"] != "Open"))

# Held awards from PROFILE
held_tracker = []
for a in PROFILE["held_awards"]:
    held_tracker.append((a, "Awarded", "N/A", "N/A", "N/A", "Yes", "—", "Currently active. Monitor renewal requirements."))

track_data = list(held_tracker)

for s in track_items:
    if s["status"] == "Open":
        default_status = "Preparing"
    elif s["status"] == "Closed" and s["your_eligible"] == "Conditional":
        default_status = "Monitor"
    else:
        default_status = "Monitor (Next Cycle)"
    priority = "High" if s["match_score"] >= 7 else "Medium" if s["match_score"] >= 4 else "Low"
    note = s["notes"][:120] + ("..." if len(s["notes"]) > 120 else "")
    track_data.append((
        s["name"], default_status, s["deadline"],
        s["link"] if str(s["link"]).startswith("http") else "",
        "", "No", priority, note,
    ))

status_fills = {
    "Awarded": PatternFill("solid", fgColor="C6EFCE"),
    "Preparing": PatternFill("solid", fgColor="FFEB9C"),
    "Applied": PatternFill("solid", fgColor="BDD7EE"),
    "Interviewing": PatternFill("solid", fgColor="DDEBF7"),
    "Monitor": PatternFill("solid", fgColor="F2F2F2"),
    "Monitor (Next Cycle)": PatternFill("solid", fgColor="F2F2F2"),
}
status_fonts = {
    "Awarded": Font(name="Arial", size=10, bold=True, color="006100"),
    "Preparing": Font(name="Arial", size=10, bold=True, color="9C6500"),
    "Monitor": Font(name="Arial", size=10, color="888888"),
    "Monitor (Next Cycle)": Font(name="Arial", size=10, color="888888"),
}

for i, row_data in enumerate(track_data):
    r = i + 2
    for c, v in enumerate(row_data, 1):
        cell = ws_tracker.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c == 4 and v and str(v).startswith("http"):
            cell.font = LINK_FONT
            cell.hyperlink = str(v)
    status_cell = ws_tracker.cell(row=r, column=2)
    st = str(row_data[1])
    if st in status_fills:
        status_cell.fill = status_fills[st]
    if st in status_fonts:
        status_cell.font = status_fonts[st]
    pri_cell = ws_tracker.cell(row=r, column=7)
    p = str(row_data[6])
    if p == "High":
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
    elif p == "Medium":
        pri_cell.font = Font(name="Arial", size=10, bold=True, color="BF8F00")

ws_tracker.auto_filter.ref = f"A1:{get_column_letter(len(track_headers))}{len(track_data)+1}"
ws_tracker.freeze_panes = "A2"
auto_width(ws_tracker, max_w=55)
ws_tracker.column_dimensions["D"].width = 40
ws_tracker.column_dimensions["E"].width = 45
ws_tracker.column_dimensions["H"].width = 55


# ═════════════════════════════════════════════════════════════════════════
# SHEET 5: RANKED PRIORITIZATION
# ═════════════════════════════════════════════════════════════════════════
ws_rank = wb.create_sheet("Ranked Prioritization")
ws_rank.sheet_properties.tabColor = "C00000"

rank_pool = [s for s in S if s["your_eligible"] in ("Eligible", "Conditional")]
rank_pool.sort(key=lambda s: (-s["match_score"], s["status"] != "Open"))

rank_headers = ["Rank", "Scholarship", "Organization", "Status", "Deadline",
                "Match Score", "Benefits", "Your Eligibility", "Why"]
for c, h in enumerate(rank_headers, 1):
    ws_rank.cell(row=1, column=c, value=h)
style_header(ws_rank, len(rank_headers))

for i, s in enumerate(rank_pool):
    r = i + 2
    why = s["notes"][:150] + ("..." if len(s["notes"]) > 150 else "")
    vals = [i + 1, s["name"], s["org"], s["status"], s["deadline"],
            s["match_score"], s["benefits"], s["your_eligible"], why]
    for c, v in enumerate(vals, 1):
        cell = ws_rank.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    rank_cell = ws_rank.cell(row=r, column=1)
    rank_cell.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    rank_cell.alignment = Alignment(horizontal="center")
    ms = ws_rank.cell(row=r, column=6)
    ms.font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ms.alignment = Alignment(horizontal="center")
    ec = ws_rank.cell(row=r, column=8)
    if s["your_eligible"] == "Eligible":
        ec.fill = ELIGIBLE_FILL
    elif s["your_eligible"] == "Conditional":
        ec.fill = COND_FILL

ws_rank.freeze_panes = "A2"
auto_width(ws_rank, max_w=55)


# ═════════════════════════════════════════════════════════════════════════
# SHEET 6: COMPARISON MATRIX
# ═════════════════════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("Comparison Matrix")
ws_comp.sheet_properties.tabColor = "7030A0"

comp_headers = ["#", "Scholarship", "Org", "Status", "Deadline", "Match",
                "Your Eligibility", "Benefits (Short)", "Exam", "Interview"]
for c, h in enumerate(comp_headers, 1):
    ws_comp.cell(row=1, column=c, value=h)
style_header(ws_comp, len(comp_headers))

for i, s in enumerate(S):
    r = i + 2
    short_ben = s["benefits"][:100] + ("..." if len(s["benefits"]) > 100 else "")
    vals = [i + 1, s["name"], s["org"], s["status"], s["deadline"],
            s["match_score"], s["your_eligible"], short_ben, s["exam"], s["interview"]]
    for c, v in enumerate(vals, 1):
        cell = ws_comp.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ec2 = ws_comp.cell(row=r, column=7)
    style_elig_cell(ec2, s["your_eligible"])
    ws_comp.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    ws_comp.cell(row=r, column=6).font = Font(name="Arial", size=10, bold=True, color="1F4E79")

ws_comp.auto_filter.ref = f"A1:{get_column_letter(len(comp_headers))}{len(S)+1}"
ws_comp.freeze_panes = "A2"
auto_width(ws_comp, max_w=50)


# ═════════════════════════════════════════════════════════════════════════
# SHEET 7: TACTICAL MANUALS
# ═════════════════════════════════════════════════════════════════════════
ws_tact = wb.create_sheet("Tactical Manuals")
ws_tact.sheet_properties.tabColor = "00B050"

tact_headers = ["Scholarship", "Match Score", "Your Eligibility", "Justification",
                "Execution Issues", "Documents to Prepare", "Strategic Tips"]
for c, h in enumerate(tact_headers, 1):
    ws_tact.cell(row=1, column=c, value=h)
style_header(ws_tact, len(tact_headers))

tacticals = [
    {
        "name": "DOST JLSS (RA 7687 & Merit)",
        "score": "10/10",
        "elig": "Eligible",
        "just": "Your 1.26–1.50 GWA far exceeds the 83% threshold. BSIT is a DOST priority course. BatStateU is a premier SUC.",
        "issues": "Strict 'no double-dipping' rule for national merit grants. May need to forfeit CHED-TDP if awarded. Qualifying exam covers advanced physics/chem/calculus outside BSIT curriculum.",
        "docs": "Certified True Copy of Grades (1st & 2nd yr), Form C (Good Moral), Form D (Program Cert), Parent BIR Form 2316 or Barangay Indigency",
        "tips": "Focus study on logical reasoning and quantitative math for JLSS exam. Monitor DOST-SEI portal for Sept 2026 qualifier announcements.",
    },
    {
        "name": "GBF STEM College Scholarship",
        "score": "9/10",
        "elig": "Eligible",
        "just": "Designed for STEM students with financial need. Direct pipeline to JG Summit tech roles. Aligns with SWE career goal.",
        "issues": "Priority given to those without other active scholarships despite allowing existing grants. Mandatory return service at Gokongwei Group post-graduation.",
        "docs": "Certified TCG (sophomore semesters), 2025-2026 registration form, utility bill, parent Certificate of Employment with compensation",
        "tips": "Build a clean digital portfolio (GitHub repos, web apps, DB scripts). During panel interview, emphasize interest in automation, cloud, or data platforms within Gokongwei Group.",
    },
    {
        "name": "Presidential Scholars (B BBM)",
        "score": "9/10",
        "elig": "Eligible",
        "just": "Province-wide Batangas rollout. 5 scholars per barangay. Santa Teresita qualifies. Fully compatible with ELAP & Provincial.",
        "issues": "Managed through local barangay leagues; timelines vary by barangay. Highly localized selection.",
        "docs": "Barangay Certificate of Residency, Barangay Indigency, Certificate of Registration (COR) from BatStateU confirming BSIT enrollment",
        "tips": "Visit Barangay Hall in Calayaan immediately. Connect with Barangay Captain. Submit certified grades + residency docs on the first day of enrollment period.",
    },
    {
        "name": "CHED Tulong Dunong (TDP)",
        "score": "9/10",
        "elig": "Eligible",
        "just": "Direct financial assistance for BatStateU students. Can be used as stipend. Compatible with existing awards.",
        "issues": "Congressional allocations are highly competitive. Cannot be combined with full TES.",
        "docs": "Wet-signature COR from BatStateU Registrar, Barangay Indigency, clear student ID copy",
        "tips": "Contact your congressional representative's staff (Santa Teresita district). Secure district application code. Submit physical docs during the designated 3-4 day enrollment window.",
    },
    {
        "name": "CHED Tertiary Education Subsidy (TES)",
        "score": "9/10",
        "elig": "Eligible",
        "just": "Up to ₱40k/yr cash subsidy under Free Higher Education law. For SUC students like BatStateU.",
        "issues": "Non-Listahanan students have lower priority. National budget limits apply.",
        "docs": "Updated COR, Certified True Copy of Grades from BatStateU Registrar",
        "tips": "Coordinate with BatStateU Scholarship Office during enrollment. Ensure registrar encoded correct GWA. Non-Listahanan students should emphasize financial need in application.",
    },
    {
        "name": "Fuji Haya Electric Scholarship",
        "score": "8/10",
        "elig": "Eligible",
        "just": "Direct BatStateU corporate partner. Targets IT/engineering. Includes OJT internships.",
        "issues": "Prohibits holding other private corporate scholarships. ELAP and Provincial LGU are ok.",
        "docs": "BatStateU internal scholarship form, TCG (sophomore yr), Dean's recommendation from College of Informatics",
        "tips": "Visit Scholarship Office at Alangilan Campus. During interview, highlight interest in industrial automation, IoT, or power system software.",
    },
    {
        "name": "BPI-DOST Innovation Awards",
        "score": "7/10",
        "elig": "Eligible",
        "just": "Prestigious competition for junior/senior applied science students. Group project format.",
        "issues": "Requires 3-member team and dean-endorsed lab-validated prototype (TRL 4). Not a traditional stipend.",
        "docs": "Contestant form, certified grades, non-disciplinary cert from Dean, 3-min video pitch",
        "tips": "Team up with classmates. Design a software solution (IoT, mobile health, agri-data platform) that solves a local problem. Make a compelling video pitch.",
    },
    {
        "name": "DICT STEM/ICT Scholarship",
        "score": "7/10",
        "elig": "Eligible",
        "just": "Targets ICT talent. BSIT directly aligns with DICT's strategic goals.",
        "issues": "Timelines and disbursements subject to national coordination. May have delays.",
        "docs": "Certified COR, student ID, freshman & sophomore grades transcript",
        "tips": "Highlight any tech certifications (AWS, Cisco, Google IT). Emphasize goal of becoming a locally-based SWE to align with DICT's digital workforce mandate.",
    },
    {
        "name": "Accenture Internship & Sponsorship",
        "score": "7/10",
        "elig": "Eligible",
        "just": "Direct industry pipeline. ABAP/SAP upskilling for IT students. Aligns with SWE goals.",
        "issues": "Career-development program, not traditional cash stipend. Focus on upskilling rather than monthly allowance.",
        "docs": "Updated technical resume, Letter of Intent, Certified TCG from BatStateU",
        "tips": "Highlight proficiency in Python, Java, or SQL. Show interest in enterprise software (SAP/ABAP, Cloud Computing) during interview.",
    },
    {
        "name": "DBP INSPIRE Scholarship",
        "score": "4/10",
        "elig": "Conditional",
        "just": "Semestral allowance for underprivileged SUC students. Good GWA match.",
        "issues": "Program details and slots managed internally via university. Requires verification with BatStateU Scholarship Office.",
        "docs": "True copy of grades, parent ITR, enrollment verification certificate",
        "tips": "Proactively visit BatStateU Scholarship Office to inquire about active DBP INSPIRE slots for the current cycle.",
    },
    {
        "name": "Huawei Seeds for the Future",
        "score": "5/10",
        "elig": "Eligible",
        "just": "High-quality tech training and certification vouchers. Excellent for upskilling.",
        "issues": "Short-term training program, not ongoing cash stipend. Focus on competitions and global networking.",
        "docs": "Resume, personal statement, digital coding portfolio",
        "tips": "Showcase strong interest in emerging tech (Cloud, AI). High GPA helps. This complements other cash stipend scholarships.",
    },
    {
        "name": "CHED CoScho",
        "score": "3/10",
        "elig": "Conditional",
        "just": "Up to ₱40k/sem if parent is a registered coconut farmer.",
        "issues": "Requires NCFRS registration. Must confirm if parent is listed.",
        "docs": "Parent NCFRS certificate, birth certificate, latest college grades",
        "tips": "Check with parents if they are registered in the National Coconut Farmers Registry. If yes, apply through CHED regional portal.",
    },
    {
        "name": "CSC Gawad PASUC",
        "score": "5/10",
        "elig": "Conditional",
        "just": "Educational subsidies at any PASUC SUC. BatStateU qualifies.",
        "issues": "Restricted to Gawad Lingkod Bayani awardees and their immediate kin.",
        "docs": "CSC nomination/award certificate, PDS form, Good Moral, enrollment verification",
        "tips": "Check if any parent/guardian is a Gawad awardee. If eligible, coordinate with CSC Central Office for endorsement.",
    },
]

for i, t in enumerate(tacticals):
    r = i + 2
    vals = [t["name"], t["score"], t["elig"], t["just"], t["issues"], t["docs"], t["tips"]]
    for c, v in enumerate(vals, 1):
        cell = ws_tact.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ec3 = ws_tact.cell(row=r, column=3)
    if t["elig"] == "Eligible":
        ec3.fill = ELIGIBLE_FILL
        ec3.font = Font(name="Arial", size=10, bold=True, color="006100")
    elif t["elig"] == "Conditional":
        ec3.fill = COND_FILL
        ec3.font = Font(name="Arial", size=10, bold=True, color="9C6500")
    ws_tact.cell(row=r, column=2).font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_tact.cell(row=r, column=2).alignment = Alignment(horizontal="center")

ws_tact.freeze_panes = "A2"
ws_tact.column_dimensions["A"].width = 38
ws_tact.column_dimensions["B"].width = 10
ws_tact.column_dimensions["C"].width = 14
ws_tact.column_dimensions["D"].width = 55
ws_tact.column_dimensions["E"].width = 50
ws_tact.column_dimensions["F"].width = 50
ws_tact.column_dimensions["G"].width = 55

# ── Save ─────────────────────────────────────────────────────────────────
output_path = r"D:\ISKORE\iskolar-tracker.xlsx"
wb.save(output_path)
print(f"Workbook saved to {output_path}")
print(f"Total scholarships: {len(S)}")
print(f"  Eligible: {eligible}")
print(f"  Conditional: {conditional}")
print(f"  Ineligible: {ineligible}")
print(f"  Open: {open_count}")
